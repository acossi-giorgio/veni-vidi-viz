import os
import unicodedata
import re
import json
import urllib.request
import pandas as pd

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
RAW = os.path.join(ROOT, "src", "datasets", "raw")
OUT = os.path.join(ROOT, "src", "datasets", "processed")
os.makedirs(OUT, exist_ok=True)

MIN_YEAR, MAX_YEAR = 2000, 2025
HISTORY_MIN_YEAR = 1995
YEAR_MAX_LIFE = 2023
YEAR_MAX_POP  = 2023
YEAR_MAX_EDU  = 2022

MIGRATION_CODE_REMAP = {
    "ABW": "DOM",
    "AIA": "DOM",
    "ATG": "DOM",
    "BES": "DOM",
    "BMU": "USA",
    "BRB": "DOM",
    "CUW": "DOM",
    "CYM": "DOM",
    "DMA": "DOM",
    "GLP": "DOM",
    "GRD": "DOM",
    "KNA": "DOM",
    "LCA": "DOM",
    "MAF": "DOM",
    "MSR": "DOM",
    "MTQ": "DOM",
    "SPM": "CAN",
    "SXM": "DOM",
    "TCA": "DOM",
    "VCT": "DOM",
    "VGB": "DOM",
    "VIR": "USA",

    "AND": "ESP",
    "FRO": "DNK",
    "GIB": "ESP",
    "IMN": "GBR",
    "LIE": "CHE",
    "MCO": "FRA",
    "MLT": "ITA",
    "SMR": "ITA",
    "VAT": "ITA",

    "BHR": "SAU",
    "HKG": "CHN",
    "MAC": "CHN",
    "MDV": "LKA",
    "SGP": "MYS",

    "COM": "MDG",
    "CPV": "SEN",
    "MUS": "MDG",
    "MYT": "MDG",
    "REU": "MDG",
    "SHN": "AGO",
    "STP": "GAB",
    "SYC": "MDG",

    "GUF": "BRA",

    "ASM": "NZL",
    "COK": "NZL",
    "FSM": "PNG",
    "GUM": "AUS",
    "KIR": "FJI",
    "MHL": "PNG",
    "MNP": "AUS",
    "NIU": "NZL",
    "NRU": "AUS",
    "PLW": "PNG",
    "PYF": "NZL",
    "TKL": "NZL",
    "TON": "FJI",
    "TUV": "FJI",
    "WLF": "FJI",
    "WSM": "FJI",
}
COUNTRY_NAME_CODE_OVERRIDES = {
    "BHS": "Bahamas",
    "BOL": "Bolivia",
    "BRN": "Brunei",
    "COD": "DR Congo",
    "COG": "Congo",
    "CZE": "Czechia",
    "EGY": "Egypt",
    "FSM": "Micronesia",
    "GBR": "United Kingdom",
    "GMB": "Gambia",
    "IRN": "Iran",
    "KOR": "South Korea",
    "LAO": "Laos",
    "MKD": "North Macedonia",
    "PRK": "North Korea",
    "RUS": "Russia",
    "SWZ": "Eswatini",
    "SYR": "Syria",
    "TZA": "Tanzania",
    "USA": "United States",
    "VEN": "Venezuela",
    "VNM": "Vietnam",
    "YEM": "Yemen",
}

AFRICA_TOPIC_CODES = [
    "DZA","AGO","BEN","BWA","BFA","BDI","CPV","CMR","CAF","TCD","COM","COG","COD","DJI","EGY","GNQ",
    "ERI","SWZ","ETH","GAB","GMB","GHA","GIN","GNB","CIV","KEN","LSO","LBR","LBY","MDG","MWI","MLI",
    "MRT","MUS","MAR","MOZ","NAM","NER","NGA","RWA","STP","SEN","SYC","SLE","SOM","ZAF","SSD","SDN",
    "TZA","TGO","TUN","UGA","ZMB","ZWE",
]

FGM_COUNTRY_ALIASES = {
    "cote d'ivoire": "CIV",
    "cote d’ivoire": "CIV",
    "united republic of tanzania": "TZA",
}

MPI_COUNTRY_ALIASES = {
    "bolivia (plurinational state of)": "BOL",
    "congo (democratic republic of the)": "COD",
    "eswatini (kingdom of)": "SWZ",
    "tanzania (united republic of)": "TZA",
}

MPI_SOUTH_AMERICA_CODES = {
    "ARG", "BOL", "BRA", "CHL", "COL", "ECU",
    "GUY", "PER", "PRY", "SUR", "URY", "VEN",
}


def simplify_official_country_name(name):
    s = str(name).strip()
    s = re.sub(r"\s*\(the territory South of 60 deg S\)\s*", "", s, flags=re.IGNORECASE)
    s = re.sub(r"\s*\(.*?\)\s*$", "", s).strip()
    if "," in s:
        s = s.split(",", 1)[0].strip()
    if s.lower().startswith("the "):
        s = s[4:].strip()
    s = re.sub(r"\s+", " ", s)
    return s

def build_simple_country_names(code_name_official):
    simple = {
        code: simplify_official_country_name(official)
        for code, official in code_name_official.items()
    }
    simple.update(COUNTRY_NAME_CODE_OVERRIDES)
    return simple

def read_country_code_reference():
    path = os.path.join(RAW, "country_codes_raw.csv")
    df = pd.read_csv(path)

    if {"alpha-3", "name", "region"}.issubset(df.columns):
        df = df.rename(columns={
            "alpha-3": "code",
            "name": "country_official",
            "region": "region",
            "sub-region": "sub_region",
            "country-code": "num",
        })

        def pick_continent(row):
            region = str(row.get("region", "")).strip()
            sub_region = str(row.get("sub_region", "")).strip()
            code = str(row.get("code", "")).strip()
            country = str(row.get("country_official", "")).strip()

            if code == "ATA" or country == "Antarctica":
                return "Oceania"
            if region == "Americas":
                return "South America" if sub_region == "South America" else "North America"
            if region in {"Africa", "Asia", "Europe", "Oceania"}:
                return region
            return region or None

        df["continent"] = df.apply(pick_continent, axis=1)
    else:
        df = df.rename(columns={
            "Three_Letter_Country_Code": "code",
            "Continent_Name":            "continent",
            "Country_Name":              "country_official",
            "Country_Number":            "num",
        })
        remap = {
            "Asia": "Asia", "Europe": "Europe", "Africa": "Africa",
            "North America": "North America", "South America": "South America",
            "Oceania": "Oceania", "Antarctica": "Oceania",
        }
        df["continent"] = df["continent"].map(remap).fillna(df["continent"])

    df = df[["code", "country_official", "continent", "num"]].dropna(subset=["code", "country_official", "continent"])
    df["code"] = df["code"].astype(str).str.strip()
    df["country_official"] = df["country_official"].astype(str).str.strip()
    df["continent"] = df["continent"].astype(str).str.strip()
    df["num"] = pd.to_numeric(df["num"], errors="coerce")
    return df

def load_mappings():
    df = read_country_code_reference()[["code", "country_official", "continent"]].copy()
    code_continent = df.set_index("code")["continent"].to_dict()
    code_name_official = df.set_index("code")["country_official"].to_dict()
    code_name_simple = build_simple_country_names(code_name_official)
    name_code = {}
    for mapping in (code_name_official, code_name_simple):
        for k, v in mapping.items():
            name_code[unicodedata.normalize("NFKD", str(v)).lower()] = k
    code_continent.setdefault("XKX", "Europe")
    code_name_simple.setdefault("XKX", "Kosovo")
    name_code.setdefault("kosovo", "XKX")

    return code_continent, code_name_simple, name_code

CODE_CONTINENT, CODE_NAME, NAME_CODE = load_mappings()

def report(name, df, error=None):
    if error:
        print(f"  [ERR] {name} -- {error}")
    else:
        print(f"  [OK]  {name:<45} {len(df):>6} rows")

def harmonize_country_columns(df):
    df = df.copy()
    if "code" in df.columns and "country" in df.columns:
        mapped = df["code"].map(CODE_NAME)
        df["country"] = mapped.fillna(df["country"]).astype(str)
    if "origin_code" in df.columns and "origin_country" in df.columns:
        mapped = df["origin_code"].map(CODE_NAME)
        df["origin_country"] = mapped.fillna(df["origin_country"]).astype(str)
    if "dest_code" in df.columns and "dest_country" in df.columns:
        mapped = df["dest_code"].map(CODE_NAME)
        df["dest_country"] = mapped.fillna(df["dest_country"]).astype(str)
    return df

def save(name, df):
    path = os.path.join(OUT, name)
    df = harmonize_country_columns(df)
    df.to_csv(path, index=False)
    report(name, df)

def harmonize_all_processed_country_names():
    for fname in os.listdir(OUT):
        if not fname.endswith(".csv"):
            continue
        path = os.path.join(OUT, fname)
        try:
            df = pd.read_csv(path)
        except Exception:
            continue
        cols = set(df.columns)
        if not ({"country", "origin_country", "dest_country"} & cols):
            continue
        df = harmonize_country_columns(df)
        df.to_csv(path, index=False)

def scope_universe_codes(scope):
    if scope == "africa_europe":
        return sorted([c for c, cont in CODE_CONTINENT.items() if cont in {"Africa", "Europe"}])
    if scope == "africa":
        return sorted([c for c, cont in CODE_CONTINENT.items() if cont == "Africa"])
    return sorted(CODE_CONTINENT.keys())

def coverage_no_data_and_incomplete(df, universe_codes, years, value_col="value"):
    if df.empty:
        return sorted(universe_codes), []
    work = df.copy()
    work = work[work["code"].isin(universe_codes)]
    work = work[work["year"].between(min(years), max(years))]
    if value_col in work.columns:
        work = work[work[value_col].notna()]
    present = {}
    for code, grp in work.groupby("code"):
        ys = set(int(y) for y in grp["year"].dropna().astype(int).tolist())
        present[code] = ys
    no_data = []
    incomplete = []
    target_years = set(int(y) for y in years)
    for code in universe_codes:
        ys = present.get(code, set())
        if not ys:
            no_data.append(code)
        elif ys != target_years:
            incomplete.append(code)
    return sorted(no_data), sorted(incomplete)

def append_missing_rows(rows, dataset_name, missing_type, codes):
    for code in sorted(set(codes)):
        rows.append({
            "dataset": dataset_name,
            "missing_type": missing_type,
            "code": code,
            "country": CODE_NAME.get(code, code),
        })

def coverage_snapshot_no_data(df, universe_codes, value_col="value"):
    if df.empty:
        return sorted(universe_codes)
    f = df.copy()
    f = f[f["code"].isin(universe_codes)]
    if value_col in f.columns:
        f = f[f[value_col].notna()]
    present = set(f["code"].dropna().astype(str).tolist())
    return sorted([c for c in universe_codes if c not in present])

def income_missing_for_child_labor_year(income_df, child_labor_df, africa_codes):
    inc_pairs = set(
        income_df[["code", "year"]]
        .dropna()
        .assign(code=lambda d: d["code"].astype(str), year=lambda d: d["year"].astype(int))
        .apply(lambda r: f"{r['code']}|{r['year']}", axis=1)
        .tolist()
    )
    cl = child_labor_df.copy()
    cl = cl[(cl["code"].isin(africa_codes)) & (cl["value"].notna())]
    latest = cl.sort_values(["code", "year"]).groupby("code", as_index=False).tail(1)
    missing = []
    for _, row in latest.iterrows():
        key = f"{str(row['code'])}|{int(row['year'])}"
        if key not in inc_pairs:
            missing.append(str(row["code"]))
    return sorted(set(missing))

def migration_origin_coverage(migration_df, origin_codes, years):
    f = migration_df.copy()
    f = f[
        (f["origin_continent"] == "Africa")
        & (f["dest_continent"] != "Africa")
        & f["year"].isin(years)
        & (f["stock"] > 0)
    ]
    target = set(int(y) for y in years)
    present_years = {
        c: set(int(y) for y in grp["year"].dropna().astype(int).tolist())
        for c, grp in f.groupby("origin_code")
    }
    no_data, incomplete = [], []
    for code in origin_codes:
        ys = present_years.get(code, set())
        if not ys:
            no_data.append(code)
        elif ys != target:
            incomplete.append(code)
    return sorted(no_data), sorted(incomplete)

def normalize_reference_year_label(value):
    text = str(value or "").strip()
    if not text or text.lower() == "nan":
        return "", None, None

    text = text.replace("–", "-").replace("—", "-").replace("/", "-")
    matches = re.findall(r"\d{2,4}", text)
    if not matches:
        return text, None, None

    start = int(matches[0])
    if start < 100:
        start += 2000 if start <= 30 else 1900

    if len(matches) == 1:
        return str(start), start, start

    end_raw = int(matches[-1])
    if end_raw < 100:
        century = (start // 100) * 100
        end = century + end_raw
        if end < start:
            end += 100
    else:
        end = end_raw

    return f"{start}-{end}", start, end

def make_missing_data_registry():
    try:
        rows = []
        ae_codes = scope_universe_codes("africa_europe")
        af_codes = scope_universe_codes("africa")

        income = pd.read_csv(os.path.join(OUT, "income.csv"))
        life = pd.read_csv(os.path.join(OUT, "life_expectancy.csv"))
        pop = pd.read_csv(os.path.join(OUT, "population.csv"))
        mpi = pd.read_csv(os.path.join(OUT, "multidimensional_poverty_index.csv"))
        spend = pd.read_csv(os.path.join(OUT, "education_spending.csv"))
        gpi = pd.read_csv(os.path.join(OUT, "gender_parity_secondary.csv"))
        oos = pd.read_csv(os.path.join(OUT, "out_of_school_rate.csv"))
        literacy = pd.read_csv(os.path.join(OUT, "youth_literacy.csv"))
        child_labor = pd.read_csv(os.path.join(OUT, "child_labor.csv"))
        child_marriage = pd.read_csv(os.path.join(OUT, "child_marriage_prevalence.csv"))
        migration = pd.read_csv(os.path.join(OUT, "migration.csv"))
        fgm = pd.read_csv(os.path.join(OUT, "fgm_quintile_prevalence.csv"))

        years_income = list(range(MIN_YEAR, MAX_YEAR + 1))
        years_life   = list(range(MIN_YEAR, YEAR_MAX_LIFE + 1))
        years_pop    = list(range(MIN_YEAR, YEAR_MAX_POP + 1))
        years_edu    = list(range(MIN_YEAR, YEAR_MAX_EDU + 1))

        # income.csv — choropleth usa tutti gli anni disponibili
        nd, inc = coverage_no_data_and_incomplete(income, ae_codes, years_income)
        append_missing_rows(rows, "income.csv", "no_data", nd)
        append_missing_rows(rows, "income.csv", "incomplete", inc)

        # life_expectancy.csv — bubble chart cap a 2023
        nd, inc = coverage_no_data_and_incomplete(life, ae_codes, years_life)
        append_missing_rows(rows, "life_expectancy.csv", "no_data", nd)
        append_missing_rows(rows, "life_expectancy.csv", "incomplete", inc)

        # population.csv — bubble chart cap a 2023
        nd, inc = coverage_no_data_and_incomplete(pop, ae_codes, years_pop)
        append_missing_rows(rows, "population.csv", "no_data", nd)
        append_missing_rows(rows, "population.csv", "incomplete", inc)

        # multidimensional_poverty_index.csv
        nd = coverage_snapshot_no_data(mpi, af_codes)
        append_missing_rows(rows, "multidimensional_poverty_index.csv", "no_data", nd)

        # education_spending.csv — chart cap a 2022
        nd, inc = coverage_no_data_and_incomplete(spend, ae_codes, years_edu)
        append_missing_rows(rows, "education_spending.csv", "no_data", nd)
        append_missing_rows(rows, "education_spending.csv", "incomplete", inc)

        # gender_parity_secondary.csv — chart usa solo ultimo anno per paese (snapshot)
        nd = coverage_snapshot_no_data(gpi, ae_codes)
        append_missing_rows(rows, "gender_parity_secondary.csv", "no_data", nd)

        # out_of_school_rate.csv — chart cap a 2022
        nd, inc = coverage_no_data_and_incomplete(oos, ae_codes, years_edu)
        append_missing_rows(rows, "out_of_school_rate.csv", "no_data", nd)
        append_missing_rows(rows, "out_of_school_rate.csv", "incomplete", inc)

        # youth_literacy.csv — chart cap a 2022
        nd, inc = coverage_no_data_and_incomplete(literacy, ae_codes, years_edu)
        append_missing_rows(rows, "youth_literacy.csv", "no_data", nd)
        append_missing_rows(rows, "youth_literacy.csv", "incomplete", inc)

        # child_labor.csv
        nd = coverage_snapshot_no_data(child_labor, af_codes)
        append_missing_rows(rows, "child_labor.csv", "no_data", nd)
        append_missing_rows(rows, "child_labor.csv", "incomplete", [])

        # income join missing for child_labor year (used by chart-4-1 merge)
        nd_inc_income = income_missing_for_child_labor_year(income, child_labor, af_codes)
        append_missing_rows(rows, "income.csv", "no_data", nd_inc_income)

        # child_marriage_prevalence.csv
        cm = child_marriage.copy()
        if "by18_pct" in cm.columns:
            cm = cm[cm["by18_pct"].notna()].rename(columns={"by18_pct": "value"})
        nd = coverage_snapshot_no_data(cm, ae_codes)
        append_missing_rows(rows, "child_marriage_prevalence.csv", "no_data", nd)
        append_missing_rows(rows, "child_marriage_prevalence.csv", "incomplete", [])

        # migration.csv
        nd, inc = migration_origin_coverage(migration, AFRICA_TOPIC_CODES, [2000, 2005, 2010, 2015, 2020])
        append_missing_rows(rows, "migration.csv", "no_data", nd)
        append_missing_rows(rows, "migration.csv", "incomplete", inc)

        # fgm_quintile_prevalence.csv
        nd = coverage_snapshot_no_data(fgm, af_codes, value_col="quintile_mean")
        append_missing_rows(rows, "fgm_quintile_prevalence.csv", "no_data", nd)
        append_missing_rows(rows, "fgm_quintile_prevalence.csv", "incomplete", [])

        registry = pd.DataFrame(
            rows,
            columns=["dataset", "missing_type", "code", "country"]
        ).drop_duplicates().sort_values(["missing_type", "dataset", "country"])
        save("missing_data_registry.csv", registry)
    except Exception as e:
        report("missing_data_registry.csv", pd.DataFrame(), str(e))

def owid_rename(df):
    return df.rename(columns={
        df.columns[0]: "country",
        df.columns[1]: "code",
        df.columns[2]: "year",
        df.columns[3]: "value",
    })

def filter_countries(df):
    df = df[df["code"].notna() & df["code"].str.len().eq(3) & (df["code"] != "")].copy()
    df["continent"] = df["code"].map(CODE_CONTINENT)
    return df[df["continent"].notna()]

def year_range(df, lo=MIN_YEAR, hi=MAX_YEAR):
    df = df.copy()
    df["year"] = df["year"].astype(int)
    return df[df["year"].between(lo, hi)]

def normalise(s):
    return unicodedata.normalize("NFKD", str(s)).lower().strip()

def make_edu_spending():
    try:
        df = pd.read_csv(os.path.join(RAW, "education_spending_raw.csv"), skiprows=4, encoding="utf-8-sig")
        df = df.rename(columns={"Country Name": "country", "Country Code": "code"})
        df = df[df["code"].str.len() == 3]
        year_cols = [c for c in df.columns if str(c).isdigit() and HISTORY_MIN_YEAR <= int(c) <= YEAR_MAX_EDU]
        melted = df.melt(id_vars=["code", "country"], value_vars=year_cols,
                         var_name="year", value_name="value")
        melted["year"] = melted["year"].astype(int)
        melted = melted.dropna(subset=["value"])
        melted["continent"] = melted["code"].map(CODE_CONTINENT)
        melted = melted[melted["continent"].notna()]
        save("education_spending.csv",
             melted[["code", "country", "continent", "year", "value"]].sort_values(["code", "year"]))
    except Exception as e:
        report("education_spending.csv", pd.DataFrame(), str(e))

def make_literacy():
    try:
        p = os.path.join(RAW, "literacy_world_bank_raw.csv")
        if not os.path.exists(p):
            raise FileNotFoundError("literacy world bank raw file not found")
        df = pd.read_csv(p, skiprows=4, encoding="utf-8-sig")
        df = df.rename(columns={"Country Name": "country", "Country Code": "code"})
        df = df[df["code"].notna() & df["code"].str.len().eq(3)]
        year_cols = [c for c in df.columns if str(c).isdigit() and HISTORY_MIN_YEAR <= int(c) <= YEAR_MAX_EDU]
        melted = df.melt(
            id_vars=["code", "country"],
            value_vars=year_cols,
            var_name="year",
            value_name="value",
        )
        melted["year"] = melted["year"].astype(int)
        melted["value"] = pd.to_numeric(melted["value"], errors="coerce")
        melted = melted.dropna(subset=["value"])
        melted["continent"] = melted["code"].map(CODE_CONTINENT)
        melted = melted[melted["continent"].notna()]
        save("youth_literacy.csv",
             melted[["code", "country", "continent", "year", "value"]].sort_values(["code", "year"]))
    except Exception as e:
        report("youth_literacy.csv", pd.DataFrame(), str(e))

def make_income():
    try:
        df = pd.read_csv(os.path.join(RAW, "income_raw.csv"))
        df = df.rename(columns={
            "Country Name": "country", "Code": "code",
            "Year": "year", "GDP_Per_Capita (USD)": "value",
        })[["code", "country", "year", "value"]]
        df = df[df["code"].notna() & df["code"].str.len().eq(3)]
        df["continent"] = df["code"].map(CODE_CONTINENT)
        df = df[df["continent"].notna()]
        df["year"] = df["year"].astype(int)
        df = df[df["year"].between(HISTORY_MIN_YEAR, MAX_YEAR)]
        df = df.dropna(subset=["value"])
        save("income.csv",
             df[["code", "country", "continent", "year", "value"]].sort_values(["code", "year"]))
    except Exception as e:
        report("income.csv", pd.DataFrame(), str(e))

def make_population():
    try:
        df = pd.read_excel(
            os.path.join(RAW, "population_raw.xlsx"),
            sheet_name="Estimates", header=16,
            usecols=[1, 5, 10, 12, 13, 14],
        )
        df.columns = ["variant", "code", "year", "pop_5_9", "pop_10_14", "pop_15_19"]
        df = df[df["variant"] == "Estimates"].copy()
        df = df[df["code"].notna() & (df["code"] != "")]
        for col in ["pop_5_9", "pop_10_14", "pop_15_19"]:
            df[col] = pd.to_numeric(df[col], errors="coerce")
        df["value"] = (
            df["pop_5_9"].fillna(0) + df["pop_10_14"].fillna(0) + df["pop_15_19"].fillna(0) * 0.6
        ) * 1000
        df["year"] = df["year"].astype(int)
        df = df[df["year"].between(HISTORY_MIN_YEAR, YEAR_MAX_POP)]
        df["continent"] = df["code"].map(CODE_CONTINENT)
        df["country"]   = df["code"].map(CODE_NAME)
        df = df[df["continent"].notna() & df["country"].notna()]
        save("population.csv",
             df[["code", "country", "continent", "year", "value"]].sort_values(["code", "year"]))
    except Exception as e:
        report("population.csv", pd.DataFrame(), str(e))

def make_child_labor():
    try:
        df = owid_rename(pd.read_csv(os.path.join(RAW, "child_labor_raw.csv")))
        df = filter_countries(df)
        df = year_range(df)
        df = df[df["value"].notna()]
        save("child_labor.csv",
             df[["code", "country", "continent", "year", "value"]].sort_values(["code", "year"]))
    except Exception as e:
        report("child_labor.csv", pd.DataFrame(), str(e))
        
def make_fgm_quintile_prevalence():
    try:
        path = os.path.join(RAW, "fgm_quintile_prevalence_raw.xlsx")
        df = pd.read_excel(path, sheet_name="Daughters FGM", header=6)

        keep_idx = [0, 7, 9, 11, 13, 15, 17]
        out = df.iloc[:, keep_idx].copy()
        out.columns = [
            "country",
            "poorest",
            "second",
            "middle",
            "fourth",
            "richest",
            "reference_year",
        ]

        out["country"] = out["country"].astype(str).str.strip()
        quintile_cols = ["poorest", "second", "middle", "fourth", "richest"]
        for col in quintile_cols:
            out[col] = pd.to_numeric(out[col], errors="coerce")

        out = out[out[quintile_cols].notna().all(axis=1)].copy()

        def country_to_code(name):
            key = normalise(name)
            code = NAME_CODE.get(key)
            if not code:
                code = FGM_COUNTRY_ALIASES.get(key)
            return code

        out["code"] = out["country"].map(country_to_code)
        out["continent"] = out["code"].map(CODE_CONTINENT)

        out = out[out["continent"] == "Africa"].copy()
        out = out[out["code"].isin(AFRICA_TOPIC_CODES)].copy()

        ref_parts = out["reference_year"].apply(
            lambda value: normalize_reference_year_label(
                str(value).strip().replace(".0", "") if pd.notna(value) else ""
            )
        )
        out["reference_year"] = ref_parts.apply(lambda item: item[0])
        out["reference_year_start"] = ref_parts.apply(lambda item: item[1])
        out["reference_year_end"] = ref_parts.apply(lambda item: item[2])
        out[quintile_cols] = out[quintile_cols].round(1)
        out["quintile_mean"] = out[quintile_cols].mean(axis=1).round(2)

        out = out[
            [
                "code",
                "country",
                "continent",
                "reference_year",
                "reference_year_start",
                "reference_year_end",
                "poorest",
                "second",
                "middle",
                "fourth",
                "richest",
                "quintile_mean",
            ]
        ].sort_values(["quintile_mean", "country"], ascending=[False, True])

        save("fgm_quintile_prevalence.csv", out)
    except Exception as e:
        report("fgm_quintile_prevalence.csv", pd.DataFrame(), str(e))

def make_out_of_school_rate():
    try:
        df = pd.read_csv(os.path.join(RAW, "out_of_school_rate_raw.csv"), skiprows=4, encoding="utf-8-sig")
        df = df.rename(columns={"Country Name": "country", "Country Code": "code"})
        df = df[df["code"].notna() & df["code"].str.len().eq(3)]
        year_cols = [c for c in df.columns if str(c).isdigit() and HISTORY_MIN_YEAR <= int(c) <= YEAR_MAX_EDU]
        melted = df.melt(
            id_vars=["code", "country"],
            value_vars=year_cols,
            var_name="year",
            value_name="value",
        )
        melted["year"] = melted["year"].astype(int)
        melted["value"] = pd.to_numeric(melted["value"], errors="coerce")
        melted = melted.dropna(subset=["value"])
        melted["continent"] = melted["code"].map(CODE_CONTINENT)
        melted = melted[melted["continent"].notna()]
        save("out_of_school_rate.csv",
             melted[["code", "country", "continent", "year", "value"]].sort_values(["code", "year"]))
    except Exception as e:
        report("out_of_school_rate.csv", pd.DataFrame(), str(e))

def make_gpi_secondary():
    try:
        df = owid_rename(pd.read_csv(os.path.join(RAW, "gender_parity_secondary_raw.csv")))
        df = filter_countries(df)
        df = year_range(df)
        df = df[df["value"].notna()]
        save("gender_parity_secondary.csv",
             df[["code", "country", "continent", "year", "value"]].sort_values(["code", "year"]))
    except Exception as e:
        report("gender_parity_secondary.csv", pd.DataFrame(), str(e))

def make_life_expectancy():
    try:
        df = pd.read_csv(os.path.join(RAW, "life_expectancy_raw.csv"))
        df = df.rename(columns={
            "Entity": "country", "Code": "code",
            "Year": "year", "Life expectancy": "value",
        })
        df = df[["code", "country", "year", "value"]].copy()
        df = filter_countries(df)
        df = year_range(df, hi=YEAR_MAX_LIFE)
        df = df[df["value"].notna()]
        save("life_expectancy.csv",
             df[["code", "country", "continent", "year", "value"]].sort_values(["code", "year"]))
    except Exception as e:
        report("life_expectancy.csv", pd.DataFrame(), str(e))

def make_migration():
    try:
        import openpyxl

        cc = read_country_code_reference()[["code", "num"]].dropna()
        cc["num"] = cc["num"].astype(int)
        num_to_code = cc.set_index("num")["code"].to_dict()
        num_to_code[729] = "SDN"

        wb = openpyxl.load_workbook(
            os.path.join(RAW, "migration_bilateral_raw.xlsx"), read_only=True)
        ws = wb["Table 1"]
        header = list(ws.iter_rows(min_row=11, max_row=11, values_only=True))[0]
        years = [v for v in header[7:14] if isinstance(v, int)]

        rows = []
        for r in ws.iter_rows(min_row=12, values_only=True):
            dest_loc = r[3]
            orig_loc = r[6]
            if not (isinstance(dest_loc, int) and 1 <= dest_loc <= 899):
                continue
            if not (isinstance(orig_loc, int) and 1 <= orig_loc <= 899):
                continue
            dest_code = num_to_code.get(dest_loc)
            orig_code = num_to_code.get(orig_loc)
            if dest_code in MIGRATION_CODE_REMAP:
                dest_code = MIGRATION_CODE_REMAP[dest_code]
            if orig_code in MIGRATION_CODE_REMAP:
                orig_code = MIGRATION_CODE_REMAP[orig_code]
            if not dest_code or not dest_code in CODE_CONTINENT:
                continue
            if not orig_code or not orig_code in CODE_CONTINENT:
                continue
            dest_name = CODE_NAME.get(dest_code, str(r[1]).strip())
            orig_name = CODE_NAME.get(orig_code, str(r[5]).strip())
            for i, yr in enumerate(years):
                val = r[7 + i]
                if val is not None and val > 0:
                    rows.append({
                        "origin_code":      orig_code,
                        "origin_country":   orig_name,
                        "origin_continent": CODE_CONTINENT[orig_code],
                        "dest_code":        dest_code,
                        "dest_country":     dest_name,
                        "dest_continent":   CODE_CONTINENT[dest_code],
                        "year":             yr,
                        "stock":            int(val),
                    })
        wb.close()

        df = pd.DataFrame(rows)
        df = df[df["year"].between(MIN_YEAR, MAX_YEAR)]
        df = (
            df.groupby(
                [
                    "origin_code", "origin_country", "origin_continent",
                    "dest_code", "dest_country", "dest_continent", "year",
                ],
                as_index=False,
            )["stock"]
            .sum()
        )
        save("migration.csv", df.sort_values(["origin_code", "dest_code", "year"]))

    except Exception as e:
        report("migration.csv", pd.DataFrame(), str(e))

def make_mpi():
    try:
        df = pd.read_excel(os.path.join(RAW, "mpi_index_raw.xlsx"), sheet_name="gMPI_Table1", header=5)
        df = df.rename(columns={
            "Country": "country",
            "2013-2024": "reference_year",
            "Value": "value",
        })
        df = df[["country", "reference_year", "value"]].copy()
        df["country"] = df["country"].astype(str).str.strip()
        df = df[df["country"].ne("") & df["country"].ne("nan")]
        df["value"] = pd.to_numeric(df["value"], errors="coerce")
        df = df[df["value"].notna()]

        code_aliases = {**NAME_CODE, **MPI_COUNTRY_ALIASES}
        df["code"] = df["country"].map(lambda name: code_aliases.get(normalise(name)))
        df["year"] = df["reference_year"].map(lambda value: normalize_reference_year_label(value)[1])
        df = df[df["code"].notna() & df["year"].notna()].copy()
        df["year"] = df["year"].astype(int)
        df["continent"] = df["code"].map(CODE_CONTINENT)
        df.loc[df["code"].isin(MPI_SOUTH_AMERICA_CODES), "continent"] = "South America"
        df = df[df["continent"].notna()]
        df["country"] = df["code"].map(CODE_NAME).fillna(df["country"])

        out = df[["code", "country", "continent", "year", "value"]].drop_duplicates()
        out = out.sort_values(["code", "year"])
        save("multidimensional_poverty_index.csv", out)
    except Exception as e:
        report("multidimensional_poverty_index.csv", pd.DataFrame(), str(e))

if __name__ == "__main__":
    make_life_expectancy()
    make_edu_spending()
    make_literacy()
    make_income()
    make_population()
    make_child_labor()
    make_fgm_quintile_prevalence()
    make_out_of_school_rate()
    make_gpi_secondary()
    make_migration()
    make_mpi()
    harmonize_all_processed_country_names()
    make_missing_data_registry()
