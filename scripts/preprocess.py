"""
Veni Vidi Viz -- Preprocessing pipeline
Schema standard: [code, country, continent, year, value]
- Tutti i CSV sono a livello paese (granularita' minima)
- Aggregazioni regionali avvengono a livello di grafico, non qui
- 'value' cambia significato per file (documentato per ciascuno)

Output -> src/datasets/processed/
Run: python scripts/preprocess.py
"""

import os
import unicodedata
import json
import urllib.request
import re
import pandas as pd

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
RAW  = os.path.join(ROOT, "src", "datasets", "raw")
OUT  = os.path.join(ROOT, "src", "datasets", "processed")
os.makedirs(OUT, exist_ok=True)

MIN_YEAR, MAX_YEAR = 2000, 2025

# Migration-map normalization:
# - world-atlas 110m omits many micro-states / dependent territories.
# - We remap them to nearby anchor countries available in the atlas so
#   all migration pairs can be rendered in map mode with a valid geometry.
# - SDN uses legacy numeric id 729 in world-atlas while ISO table often reports 736.
#   Keep code SDN, but treat it as atlas-aligned via alias in QA checks.
MIGRATION_CODE_REMAP = {
    # North America / Caribbean micro-states and territories -> Dominican Republic / US / Canada
    "ABW": "DOM",
    "AIA": "DOM",
    "ATG": "DOM",
    "BES": "DOM",
    "BMU": "USA",
    "BRB": "DOM",
    "CUW": "DOM",
    "CYM": "DOM",
    "DMA": "DOM",
    "GLP": "DOM",
    "GRD": "DOM",
    "KNA": "DOM",
    "LCA": "DOM",
    "MAF": "DOM",
    "MSR": "DOM",
    "MTQ": "DOM",
    "SPM": "CAN",
    "SXM": "DOM",
    "TCA": "DOM",
    "VCT": "DOM",
    "VGB": "DOM",
    "VIR": "USA",

    # Europe micro-states / territories
    "AND": "ESP",
    "FRO": "DNK",
    "GIB": "ESP",
    "IMN": "GBR",
    "LIE": "CHE",
    "MCO": "FRA",
    "MLT": "ITA",
    "SMR": "ITA",
    "VAT": "ITA",

    # Asia micro-states / special regions
    "BHR": "SAU",
    "HKG": "CHN",
    "MAC": "CHN",
    "MDV": "LKA",
    "SGP": "MYS",

    # Africa islands / territories
    "COM": "MDG",
    "CPV": "SEN",
    "MUS": "MDG",
    "MYT": "MDG",
    "REU": "MDG",
    "SHN": "AGO",
    "STP": "GAB",
    "SYC": "MDG",

    # South America territories
    "GUF": "BRA",

    # Oceania micro-states / territories
    "ASM": "NZL",
    "COK": "NZL",
    "FSM": "PNG",
    "GUM": "AUS",
    "KIR": "FJI",
    "MHL": "PNG",
    "MNP": "AUS",
    "NIU": "NZL",
    "NRU": "AUS",
    "PLW": "PNG",
    "PYF": "NZL",
    "TKL": "NZL",
    "TON": "FJI",
    "TUV": "FJI",
    "WLF": "FJI",
    "WSM": "FJI",
}
MIGRATION_NUM_TO_CODE_ALIAS = {
    729: "SDN",
}
MIGRATION_ATLAS_NUM_ALIAS = {
    "SDN": 729,
}

COUNTRY_NAME_CODE_OVERRIDES = {
    "BHS": "Bahamas",
    "BOL": "Bolivia",
    "BRN": "Brunei",
    "COD": "DR Congo",
    "COG": "Congo",
    "CZE": "Czechia",
    "EGY": "Egypt",
    "FSM": "Micronesia",
    "GBR": "United Kingdom",
    "GMB": "Gambia",
    "IRN": "Iran",
    "KOR": "South Korea",
    "LAO": "Laos",
    "MKD": "North Macedonia",
    "PRK": "North Korea",
    "RUS": "Russia",
    "SWZ": "Eswatini",
    "SYR": "Syria",
    "TZA": "Tanzania",
    "USA": "United States",
    "VEN": "Venezuela",
    "VNM": "Vietnam",
    "YEM": "Yemen",
}

AFRICA_TOPIC_CODES = [
    "DZA","AGO","BEN","BWA","BFA","BDI","CPV","CMR","CAF","TCD","COM","COG","COD","DJI","EGY","GNQ",
    "ERI","SWZ","ETH","GAB","GMB","GHA","GIN","GNB","CIV","KEN","LSO","LBR","LBY","MDG","MWI","MLI",
    "MRT","MUS","MAR","MOZ","NAM","NER","NGA","RWA","STP","SEN","SYC","SLE","SOM","ZAF","SSD","SDN",
    "TZA","TGO","TUN","UGA","ZMB","ZWE",
]


# ── Mappings ──────────────────────────────────────────────────────────────────
def simplify_official_country_name(name):
    s = str(name).strip()
    s = re.sub(r"\s*\(the territory South of 60 deg S\)\s*", "", s, flags=re.IGNORECASE)
    s = re.sub(r"\s*\(.*?\)\s*$", "", s).strip()
    if "," in s:
        s = s.split(",", 1)[0].strip()
    if s.lower().startswith("the "):
        s = s[4:].strip()
    s = re.sub(r"\s+", " ", s)
    return s


def build_simple_country_names(code_name_official):
    simple = {
        code: simplify_official_country_name(official)
        for code, official in code_name_official.items()
    }
    simple.update(COUNTRY_NAME_CODE_OVERRIDES)
    return simple


def load_mappings():
    path = os.path.join(RAW, "country-code.csv")
    df = pd.read_csv(path).rename(columns={
        "Three_Letter_Country_Code": "code",
        "Continent_Name":            "continent",
        "Country_Name":              "country_official",
    })[["code", "country_official", "continent"]].dropna()
    remap = {
        "Asia": "Asia", "Europe": "Europe", "Africa": "Africa",
        "North America": "North America", "South America": "South America",
        "Oceania": "Oceania", "Antarctica": "Oceania",
    }
    df["continent"] = df["continent"].map(remap).fillna(df["continent"])
    code_continent = df.set_index("code")["continent"].to_dict()
    code_name_official = df.set_index("code")["country_official"].to_dict()
    code_name_simple = build_simple_country_names(code_name_official)
    # name -> code for migration matching (normalised lowercase), include both variants
    name_code = {}
    for mapping in (code_name_official, code_name_simple):
        for k, v in mapping.items():
            name_code[unicodedata.normalize("NFKD", str(v)).lower()] = k
    # Kosovo (XKX) appears in some source datasets but is absent from country-code.csv.
    code_continent.setdefault("XKX", "Europe")
    code_name_simple.setdefault("XKX", "Kosovo")
    name_code.setdefault("kosovo", "XKX")

    return code_continent, code_name_simple, name_code

CODE_CONTINENT, CODE_NAME, NAME_CODE = load_mappings()


# ── Helpers ───────────────────────────────────────────────────────────────────
def report(name, df, error=None):
    if error:
        print(f"  [ERR] {name} -- {error}")
    else:
        print(f"  [OK]  {name:<45} {len(df):>6} rows")


def harmonize_country_columns(df):
    df = df.copy()
    if "code" in df.columns and "country" in df.columns:
        mapped = df["code"].map(CODE_NAME)
        df["country"] = mapped.fillna(df["country"]).astype(str)
    if "origin_code" in df.columns and "origin_country" in df.columns:
        mapped = df["origin_code"].map(CODE_NAME)
        df["origin_country"] = mapped.fillna(df["origin_country"]).astype(str)
    if "dest_code" in df.columns and "dest_country" in df.columns:
        mapped = df["dest_code"].map(CODE_NAME)
        df["dest_country"] = mapped.fillna(df["dest_country"]).astype(str)
    return df

def save(name, df):
    path = os.path.join(OUT, name)
    df = harmonize_country_columns(df)
    df.to_csv(path, index=False)
    report(name, df)


def harmonize_all_processed_country_names():
    for fname in os.listdir(OUT):
        if not fname.endswith(".csv"):
            continue
        path = os.path.join(OUT, fname)
        try:
            df = pd.read_csv(path)
        except Exception:
            continue
        cols = set(df.columns)
        if not ({"country", "origin_country", "dest_country"} & cols):
            continue
        df = harmonize_country_columns(df)
        df.to_csv(path, index=False)


def _scope_universe_codes(scope):
    if scope == "africa_europe":
        return sorted([c for c, cont in CODE_CONTINENT.items() if cont in {"Africa", "Europe"}])
    if scope == "africa":
        return sorted([c for c, cont in CODE_CONTINENT.items() if cont == "Africa"])
    return sorted(CODE_CONTINENT.keys())


def _coverage_no_data_and_incomplete(df, universe_codes, years, value_col="value"):
    if df.empty:
        return sorted(universe_codes), []
    work = df.copy()
    work = work[work["code"].isin(universe_codes)]
    work = work[work["year"].between(min(years), max(years))]
    if value_col in work.columns:
        work = work[work[value_col].notna()]
    present = {}
    for code, grp in work.groupby("code"):
        ys = set(int(y) for y in grp["year"].dropna().astype(int).tolist())
        present[code] = ys
    no_data = []
    incomplete = []
    target_years = set(int(y) for y in years)
    for code in universe_codes:
        ys = present.get(code, set())
        if not ys:
            no_data.append(code)
        elif ys != target_years:
            incomplete.append(code)
    return sorted(no_data), sorted(incomplete)


def _codes_to_names(codes):
    out = []
    for c in codes:
        out.append(CODE_NAME.get(c, c))
    return sorted(out)


def _names_list(names):
    return "; ".join(names) if names else "nessuno."


def _append_missing_rows(rows, dataset_name, missing_type, codes):
    for code in sorted(set(codes)):
        rows.append({
            "dataset": dataset_name,
            "missing_type": missing_type,
            "code": code,
            "country": CODE_NAME.get(code, code),
        })


def _coverage_snapshot_no_data(df, universe_codes, value_col="value"):
    if df.empty:
        return sorted(universe_codes)
    f = df.copy()
    f = f[f["code"].isin(universe_codes)]
    if value_col in f.columns:
        f = f[f[value_col].notna()]
    present = set(f["code"].dropna().astype(str).tolist())
    return sorted([c for c in universe_codes if c not in present])


def _income_missing_for_child_labor_year(income_df, child_labor_df, africa_codes):
    inc_pairs = set(
        income_df[["code", "year"]]
        .dropna()
        .assign(code=lambda d: d["code"].astype(str), year=lambda d: d["year"].astype(int))
        .apply(lambda r: f"{r['code']}|{r['year']}", axis=1)
        .tolist()
    )
    cl = child_labor_df.copy()
    cl = cl[(cl["code"].isin(africa_codes)) & (cl["value"].notna())]
    latest = cl.sort_values(["code", "year"]).groupby("code", as_index=False).tail(1)
    missing = []
    for _, row in latest.iterrows():
        key = f"{str(row['code'])}|{int(row['year'])}"
        if key not in inc_pairs:
            missing.append(str(row["code"]))
    return sorted(set(missing))


def _migration_origin_coverage(migration_df, origin_codes, years):
    f = migration_df.copy()
    f = f[
        (f["origin_continent"] == "Africa")
        & (f["dest_continent"] != "Africa")
        & f["year"].isin(years)
        & (f["stock"] > 0)
    ]
    target = set(int(y) for y in years)
    present_years = {
        c: set(int(y) for y in grp["year"].dropna().astype(int).tolist())
        for c, grp in f.groupby("origin_code")
    }
    no_data, incomplete = [], []
    for code in origin_codes:
        ys = present_years.get(code, set())
        if not ys:
            no_data.append(code)
        elif ys != target:
            incomplete.append(code)
    return sorted(no_data), sorted(incomplete)


def make_missing_data_registry():
    try:
        rows = []
        ae_codes = _scope_universe_codes("africa_europe")
        af_codes = _scope_universe_codes("africa")

        income = pd.read_csv(os.path.join(OUT, "income.csv"))
        life = pd.read_csv(os.path.join(OUT, "life_expectancy.csv"))
        pop = pd.read_csv(os.path.join(OUT, "population.csv"))
        mpi = pd.read_csv(os.path.join(OUT, "mpi.csv"))
        spend = pd.read_csv(os.path.join(OUT, "edu_spending.csv"))
        gpi = pd.read_csv(os.path.join(OUT, "gpi_secondary.csv"))
        oos = pd.read_csv(os.path.join(OUT, "out_of_school.csv"))
        literacy = pd.read_csv(os.path.join(OUT, "literacy.csv"))
        child_labor = pd.read_csv(os.path.join(OUT, "child_labor.csv"))
        child_marriage = pd.read_csv(os.path.join(OUT, "child_marriage_cmmm.csv"))
        maternal = pd.read_csv(os.path.join(OUT, "maternal_mortality.csv"))
        child_mort = pd.read_csv(os.path.join(OUT, "child_mortality.csv"))
        migration = pd.read_csv(os.path.join(OUT, "migration.csv"))

        # income.csv
        nd, inc = _coverage_no_data_and_incomplete(income, ae_codes, list(range(2000, 2025)))
        _append_missing_rows(rows, "income.csv", "no_data", nd)
        _append_missing_rows(rows, "income.csv", "incomplete", inc)

        # life_expectancy.csv
        nd, inc = _coverage_no_data_and_incomplete(life, ae_codes, list(range(2000, 2024)))
        _append_missing_rows(rows, "life_expectancy.csv", "no_data", nd)
        _append_missing_rows(rows, "life_expectancy.csv", "incomplete", inc)

        # population.csv
        nd, inc = _coverage_no_data_and_incomplete(pop, ae_codes, list(range(2000, 2024)))
        _append_missing_rows(rows, "population.csv", "no_data", nd)
        _append_missing_rows(rows, "population.csv", "incomplete", inc)

        # mpi.csv (snapshot)
        nd = _coverage_snapshot_no_data(mpi, af_codes)
        _append_missing_rows(rows, "mpi.csv", "no_data", nd)

        # edu_spending.csv
        years_edu = list(range(2000, 2023))
        nd, inc = _coverage_no_data_and_incomplete(spend, ae_codes, years_edu)
        _append_missing_rows(rows, "edu_spending.csv", "no_data", nd)
        _append_missing_rows(rows, "edu_spending.csv", "incomplete", inc)

        # gpi_secondary.csv
        years_32 = list(range(2000, 2025))
        nd, inc = _coverage_no_data_and_incomplete(gpi, ae_codes, years_32)
        _append_missing_rows(rows, "gpi_secondary.csv", "no_data", nd)
        _append_missing_rows(rows, "gpi_secondary.csv", "incomplete", inc)

        # out_of_school.csv
        nd, inc = _coverage_no_data_and_incomplete(oos, ae_codes, years_32)
        _append_missing_rows(rows, "out_of_school.csv", "no_data", nd)
        _append_missing_rows(rows, "out_of_school.csv", "incomplete", inc)

        # literacy.csv
        nd, inc = _coverage_no_data_and_incomplete(literacy, ae_codes, years_edu)
        _append_missing_rows(rows, "literacy.csv", "no_data", nd)
        _append_missing_rows(rows, "literacy.csv", "incomplete", inc)

        # child_labor.csv (snapshot)
        nd = _coverage_snapshot_no_data(child_labor, af_codes)
        _append_missing_rows(rows, "child_labor.csv", "no_data", nd)
        _append_missing_rows(rows, "child_labor.csv", "incomplete", [])

        # income join missing for child_labor year (used by chart-4-1 merge)
        nd_inc_income = _income_missing_for_child_labor_year(income, child_labor, af_codes)
        _append_missing_rows(rows, "income.csv", "no_data", nd_inc_income)

        # child_marriage_cmmm.csv (snapshot)
        cm = child_marriage.copy()
        if "by18_pct" in cm.columns:
            cm = cm[cm["by18_pct"].notna()].rename(columns={"by18_pct": "value"})
        nd = _coverage_snapshot_no_data(cm, ae_codes)
        _append_missing_rows(rows, "child_marriage_cmmm.csv", "no_data", nd)
        _append_missing_rows(rows, "child_marriage_cmmm.csv", "incomplete", [])

        # maternal_mortality.csv + child_mortality.csv
        years_43 = list(range(2000, 2024))
        nd, inc = _coverage_no_data_and_incomplete(maternal, ae_codes, years_43)
        _append_missing_rows(rows, "maternal_mortality.csv", "no_data", nd)
        _append_missing_rows(rows, "maternal_mortality.csv", "incomplete", inc)
        nd, inc = _coverage_no_data_and_incomplete(child_mort, ae_codes, years_43)
        _append_missing_rows(rows, "child_mortality.csv", "no_data", nd)
        _append_missing_rows(rows, "child_mortality.csv", "incomplete", inc)

        # migration.csv
        nd, inc = _migration_origin_coverage(migration, AFRICA_TOPIC_CODES, [2000, 2005, 2010, 2015, 2020])
        _append_missing_rows(rows, "migration.csv", "no_data", nd)
        _append_missing_rows(rows, "migration.csv", "incomplete", inc)

        registry = pd.DataFrame(
            rows,
            columns=["dataset", "missing_type", "code", "country"]
        ).drop_duplicates().sort_values(["missing_type", "dataset", "country"])
        save("missing_data.csv", registry)
    except Exception as e:
        report("missing_data.csv", pd.DataFrame(), str(e))

def owid_rename(df):
    """Rename standard OWID columns: Entity->country, Code->code, Year->year, last->value."""
    return df.rename(columns={
        df.columns[0]: "country",
        df.columns[1]: "code",
        df.columns[2]: "year",
        df.columns[3]: "value",
    })

def filter_countries(df):
    """Keep only real country rows (3-char ISO code, has continent mapping)."""
    df = df[df["code"].notna() & df["code"].str.len().eq(3) & (df["code"] != "")].copy()
    df["continent"] = df["code"].map(CODE_CONTINENT)
    return df[df["continent"].notna()]

def year_range(df, lo=MIN_YEAR, hi=MAX_YEAR):
    df = df.copy()
    df["year"] = df["year"].astype(int)
    return df[df["year"].between(lo, hi)]

def normalise(s):
    return unicodedata.normalize("NFKD", str(s)).lower().strip()


# ── edu_spending.csv ──────────────────────────────────────────────────────────
# value = Government expenditure on education, total (% of GDP)
# Fonte: World Bank SE.XPD.TOTL.GD.ZS
def make_edu_spending():
    try:
        df = pd.read_csv(os.path.join(RAW, "edu_spending.csv"), skiprows=4, encoding="utf-8-sig")
        df = df.rename(columns={"Country Name": "country", "Country Code": "code"})
        df = df[df["code"].str.len() == 3]
        year_cols = [c for c in df.columns if str(c).isdigit() and MIN_YEAR <= int(c) <= MAX_YEAR]
        melted = df.melt(id_vars=["code", "country"], value_vars=year_cols,
                         var_name="year", value_name="value")
        melted["year"] = melted["year"].astype(int)
        melted = melted.dropna(subset=["value"])
        melted["continent"] = melted["code"].map(CODE_CONTINENT)
        melted = melted[melted["continent"].notna()]
        save("edu_spending.csv",
             melted[["code", "country", "continent", "year", "value"]].sort_values(["code", "year"]))
    except Exception as e:
        report("edu_spending.csv", pd.DataFrame(), str(e))


# ── edu_completion.csv ────────────────────────────────────────────────────────
# value = Completion rate, upper secondary education, both sexes (%)
# Fonte: Our World in Data / SDG 4.1.2
def make_edu_completion():
    try:
        df = owid_rename(pd.read_csv(os.path.join(RAW, "edu_completion.csv")))
        df = filter_countries(df)
        df = year_range(df)
        df = df[df["value"].notna()]
        save("edu_completion.csv",
             df[["code", "country", "continent", "year", "value"]].sort_values(["code", "year"]))
    except Exception as e:
        report("edu_completion.csv", pd.DataFrame(), str(e))


# ── literacy.csv ──────────────────────────────────────────────────────────────
# value = Literacy rate, adult total (% of people ages 15+)
# Fonte: Our World in Data / UNESCO UIS
def make_literacy():
    try:
        for fname in ("literacy_raw.csv", "literacy.csv"):
            p = os.path.join(RAW, fname)
            if os.path.exists(p):
                df = pd.read_csv(p)
                break
        else:
            raise FileNotFoundError("literacy raw file not found")
        df = owid_rename(df)
        df = filter_countries(df)
        df = year_range(df)
        df = df[df["value"].notna()]
        save("literacy.csv",
             df[["code", "country", "continent", "year", "value"]].sort_values(["code", "year"]))
    except Exception as e:
        report("literacy.csv", pd.DataFrame(), str(e))


# ── income.csv ────────────────────────────────────────────────────────────────
# value = GDP per capita (USD, current prices)
# Fonte: World Bank / custom CSV
def make_income():
    try:
        df = pd.read_csv(os.path.join(RAW, "gdp_per_capita.csv"))
        df = df.rename(columns={
            "Country Name": "country", "Code": "code",
            "Year": "year", "GDP_Per_Capita (USD)": "value",
        })[["code", "country", "year", "value"]]
        df = df[df["code"].notna() & df["code"].str.len().eq(3)]
        df["continent"] = df["code"].map(CODE_CONTINENT)
        df = df[df["continent"].notna()]
        df["year"] = df["year"].astype(int)
        df = df[df["year"].between(MIN_YEAR, MAX_YEAR)]
        df = df.dropna(subset=["value"])
        save("income.csv",
             df[["code", "country", "continent", "year", "value"]].sort_values(["code", "year"]))
    except Exception as e:
        report("income.csv", pd.DataFrame(), str(e))


# ── population.csv ────────────────────────────────────────────────────────────
# value = Population ages 5-17 (estimated)
# Fonte: UN World Population Prospects 2024
def make_population():
    try:
        df = pd.read_excel(
            os.path.join(RAW, "child_population_raw.xlsx"),
            sheet_name="Estimates", header=16,
            usecols=[1, 5, 10, 12, 13, 14],
        )
        df.columns = ["variant", "code", "year", "pop_5_9", "pop_10_14", "pop_15_19"]
        df = df[df["variant"] == "Estimates"].copy()
        df = df[df["code"].notna() & (df["code"] != "")]
        df["value"] = (
            df["pop_5_9"].fillna(0) + df["pop_10_14"].fillna(0) + df["pop_15_19"].fillna(0) * 0.6
        ) * 1000
        df["year"] = df["year"].astype(int)
        df = df[df["year"].between(MIN_YEAR, MAX_YEAR)]
        df["continent"] = df["code"].map(CODE_CONTINENT)
        df["country"]   = df["code"].map(CODE_NAME)
        df = df[df["continent"].notna() & df["country"].notna()]
        save("population.csv",
             df[["code", "country", "continent", "year", "value"]].sort_values(["code", "year"]))
    except Exception as e:
        report("population.csv", pd.DataFrame(), str(e))



# ── child_labor.csv ───────────────────────────────────────────────────────────
# value = Share of children engaged in economic activity, ages 5-17 (%)
# Fonte: Our World in Data / ILO-UNICEF
def make_child_labor():
    try:
        df = owid_rename(pd.read_csv(os.path.join(RAW, "child_labor_raw.csv")))
        df = filter_countries(df)
        df = year_range(df)
        df = df[df["value"].notna()]
        save("child_labor.csv",
             df[["code", "country", "continent", "year", "value"]].sort_values(["code", "year"]))
    except Exception as e:
        report("child_labor.csv", pd.DataFrame(), str(e))


# ── child_marriage.csv ────────────────────────────────────────────────────────
# value = Women first married by age 18 (% of women ages 20-24)
# Fonte: Our World in Data / UNICEF
def make_child_marriage():
    try:
        df = pd.read_csv(os.path.join(RAW, "child_marriage_raw.csv"))
        df = df.rename(columns={
            df.columns[0]: "country", df.columns[1]: "code",
            df.columns[2]: "year",    df.columns[3]: "value",
        })
        df = filter_countries(df)
        df = year_range(df)
        df = df[df["value"].notna()]
        save("child_marriage.csv",
             df[["code", "country", "continent", "year", "value"]].sort_values(["code", "year"]))
    except Exception as e:
        report("child_marriage.csv", pd.DataFrame(), str(e))


# ── out_of_school.csv ────────────────────────────────────────────────────────
# value = Children out of primary school, both sexes (absolute number)
# Fonte: Our World in Data / UNESCO UIS
def make_out_of_school():
    try:
        df = pd.read_csv(os.path.join(RAW, "out_of_school_raw.csv"))
        df = df.rename(columns={
            df.columns[0]: "country",
            df.columns[1]: "code",
            df.columns[2]: "year",
            df.columns[3]: "value",
        })
        df = filter_countries(df)
        df = year_range(df)
        df = df[df["value"].notna()]
        save("out_of_school.csv",
             df[["code", "country", "continent", "year", "value"]].sort_values(["code", "year"]))
    except Exception as e:
        report("out_of_school.csv", pd.DataFrame(), str(e))


# ── poverty.csv ──────────────────────────────────────────────────────────────
# value = Share of population living on less than $2.15/day (%)
# Fonte: Our World in Data / World Bank PIP
def make_poverty():
    try:
        df = owid_rename(pd.read_csv(os.path.join(RAW, "poverty_raw.csv")))
        df = filter_countries(df)
        df = year_range(df)
        df = df[df["value"].notna()]
        save("poverty.csv",
             df[["code", "country", "continent", "year", "value"]].sort_values(["code", "year"]))
    except Exception as e:
        report("poverty.csv", pd.DataFrame(), str(e))


# ── gini.csv ──────────────────────────────────────────────────────────────────
# value = Gini coefficient (0=perfect equality, 1=perfect inequality)
# Fonte: Our World in Data / World Bank PIP
def make_gini():
    try:
        df = owid_rename(pd.read_csv(os.path.join(RAW, "gini_raw.csv")))
        df = filter_countries(df)
        df = year_range(df)
        df = df[df["value"].notna()]
        save("gini.csv",
             df[["code", "country", "continent", "year", "value"]].sort_values(["code", "year"]))
    except Exception as e:
        report("gini.csv", pd.DataFrame(), str(e))


# ── gpi_secondary.csv ─────────────────────────────────────────────────────────
# value = Adjusted gender parity index, lower secondary net enrollment
#         (1 = parity, <1 = boys favored, >1 = girls favored)
# Fonte: Our World in Data / UNESCO UIS
def make_gpi_secondary():
    try:
        df = owid_rename(pd.read_csv(os.path.join(RAW, "gpi_secondary_raw.csv")))
        df = filter_countries(df)
        df = year_range(df)
        df = df[df["value"].notna()]
        save("gpi_secondary.csv",
             df[["code", "country", "continent", "year", "value"]].sort_values(["code", "year"]))
    except Exception as e:
        report("gpi_secondary.csv", pd.DataFrame(), str(e))


# ── child_mortality.csv ───────────────────────────────────────────────────────
# value = Under-5 mortality rate (deaths per 1000 live births)
# Fonte: Our World in Data / UN IGME + Gapminder
def make_child_mortality():
    try:
        df = owid_rename(pd.read_csv(os.path.join(RAW, "child_mortality_raw.csv")))
        df = filter_countries(df)
        df = year_range(df)
        df = df[df["value"].notna()]
        save("child_mortality.csv",
             df[["code", "country", "continent", "year", "value"]].sort_values(["code", "year"]))
    except Exception as e:
        report("child_mortality.csv", pd.DataFrame(), str(e))


# ── maternal_mortality.csv ────────────────────────────────────────────────────
# value = Maternal mortality ratio (deaths per 100,000 live births)
# Fonte: Our World in Data / WHO GHO
def make_maternal_mortality():
    try:
        df = owid_rename(pd.read_csv(os.path.join(RAW, "maternal_mortality_raw.csv")))
        df = filter_countries(df)
        df = year_range(df)
        df = df[df["value"].notna()]
        save("maternal_mortality.csv",
             df[["code", "country", "continent", "year", "value"]].sort_values(["code", "year"]))
    except Exception as e:
        report("maternal_mortality.csv", pd.DataFrame(), str(e))


# ── life_expectancy.csv ───────────────────────────────────────────────────────
# value = Life expectancy at birth (years)
# Fonte: Our World in Data / UN WPP
def make_life_expectancy():
    try:
        df = pd.read_csv(os.path.join(RAW, "life_expectancy.csv"))
        # Columns: Entity, Code, Year, Life expectancy, Continent
        df = df.rename(columns={
            "Entity": "country", "Code": "code",
            "Year": "year", "Life expectancy": "value",
        })
        df = df[["code", "country", "year", "value"]].copy()
        df = filter_countries(df)
        df = year_range(df)
        df = df[df["value"].notna()]
        save("life_expectancy.csv",
             df[["code", "country", "continent", "year", "value"]].sort_values(["code", "year"]))
    except Exception as e:
        report("life_expectancy.csv", pd.DataFrame(), str(e))


# ── remittances.csv ───────────────────────────────────────────────────────────
# value = Personal remittances received (% of GDP)
# Fonte: Our World in Data / World Bank (IMF + OECD)
def make_remittances():
    try:
        df = owid_rename(pd.read_csv(os.path.join(RAW, "remittances_raw.csv")))
        df = filter_countries(df)
        df = year_range(df)
        df = df[df["value"].notna()]
        save("remittances.csv",
             df[["code", "country", "continent", "year", "value"]].sort_values(["code", "year"]))
    except Exception as e:
        report("remittances.csv", pd.DataFrame(), str(e))


# ── migration.csv ─────────────────────────────────────────────────────────────
# Flussi bilaterali: stock di migranti per coppia (origine, destinazione)
# Schema: origin_code, origin_country, origin_continent,
#         dest_code, dest_country, dest_continent, year, stock
# Fonte: UN DESA 2020 International Migrant Stock — Destination and Origin, Table 1
# Anni disponibili: 1990, 1995, 2000, 2005, 2010, 2015, 2020 (quinquennali)
def make_migration():
    try:
        import openpyxl

        # Build numeric ISO code -> ISO3 + continent mapping
        cc = pd.read_csv(os.path.join(RAW, "country-code.csv")).rename(columns={
            "Three_Letter_Country_Code": "code",
            "Country_Number": "num",
        })[["code", "num"]].dropna()
        cc["num"] = cc["num"].astype(int)
        num_to_code = cc.set_index("num")["code"].to_dict()
        num_to_code.update(MIGRATION_NUM_TO_CODE_ALIAS)

        wb = openpyxl.load_workbook(
            os.path.join(RAW, "migration_bilateral_raw.xlsx"), read_only=True)
        ws = wb["Table 1"]
        header = list(ws.iter_rows(min_row=11, max_row=11, values_only=True))[0]
        # years = first 7 year values after col 7 (both sexes combined)
        years = [v for v in header[7:14] if isinstance(v, int)]

        rows = []
        for r in ws.iter_rows(min_row=12, values_only=True):
            dest_loc = r[3]
            orig_loc = r[6]
            # keep only country-level pairs (ISO numeric 1-899)
            if not (isinstance(dest_loc, int) and 1 <= dest_loc <= 899):
                continue
            if not (isinstance(orig_loc, int) and 1 <= orig_loc <= 899):
                continue
            dest_code = num_to_code.get(dest_loc)
            orig_code = num_to_code.get(orig_loc)
            if dest_code in MIGRATION_CODE_REMAP:
                dest_code = MIGRATION_CODE_REMAP[dest_code]
            if orig_code in MIGRATION_CODE_REMAP:
                orig_code = MIGRATION_CODE_REMAP[orig_code]
            if not dest_code or not dest_code in CODE_CONTINENT:
                continue
            if not orig_code or not orig_code in CODE_CONTINENT:
                continue
            dest_name = CODE_NAME.get(dest_code, str(r[1]).strip())
            orig_name = CODE_NAME.get(orig_code, str(r[5]).strip())
            for i, yr in enumerate(years):
                val = r[7 + i]
                if val is not None and val > 0:
                    rows.append({
                        "origin_code":      orig_code,
                        "origin_country":   orig_name,
                        "origin_continent": CODE_CONTINENT[orig_code],
                        "dest_code":        dest_code,
                        "dest_country":     dest_name,
                        "dest_continent":   CODE_CONTINENT[dest_code],
                        "year":             yr,
                        "stock":            int(val),
                    })
        wb.close()

        df = pd.DataFrame(rows)
        df = df[df["year"].between(MIN_YEAR, MAX_YEAR)]
        # Merge potential collisions introduced by code remap (e.g. MAF -> FRA).
        df = (
            df.groupby(
                [
                    "origin_code", "origin_country", "origin_continent",
                    "dest_code", "dest_country", "dest_continent", "year",
                ],
                as_index=False,
            )["stock"]
            .sum()
        )
        save("migration.csv", df.sort_values(["origin_code", "dest_code", "year"]))

        # QA artifact: codes present in migration but not mappable in world-atlas 110m.
        # Uses numeric-code aliases where atlas id differs from ISO table (SDN).
        atlas = json.load(urllib.request.urlopen("https://cdn.jsdelivr.net/npm/world-atlas@2.0.2/countries-110m.json"))
        geoms = atlas.get("objects", {}).get("countries", {}).get("geometries", [])
        atlas_ids = set(int(g["id"]) for g in geoms if str(g.get("id", "")).isdigit())
        code_num = cc.set_index("code")["num"].to_dict()

        all_codes = set(df["origin_code"]).union(set(df["dest_code"]))
        unmapped = []
        for code in sorted(all_codes):
            num = code_num.get(code)
            if code in MIGRATION_ATLAS_NUM_ALIAS:
                num = MIGRATION_ATLAS_NUM_ALIAS[code]
            if num is None or int(num) not in atlas_ids:
                unmapped.append({
                    "code": code,
                    "country": CODE_NAME.get(code, code),
                    "country_number": code_num.get(code),
                })

        qa = pd.DataFrame(unmapped, columns=["code", "country", "country_number"])
        save("migration_unmapped_codes.csv", qa)
    except Exception as e:
        report("migration.csv", pd.DataFrame(), str(e))


# ── mpi.csv ───────────────────────────────────────────────────────────────────
# value = MPI score (0–1), national level only
# Fonte: OPHI Global MPI via HDX
def make_mpi():
    try:
        df = pd.read_csv(os.path.join(RAW, "multidimensional-poverty-index-mpi.csv"))
        df = df.rename(columns={
            "Code":                                    "code",
            "Entity":                                  "country",
            "Year":                                    "year",
            "Multidimensional Poverty Index (MPI)":    "value",
            "World region according to OWID":          "continent",
        })
        df = df[df["code"].notna() & (df["code"].str.strip() != "")]
        df = df[df["value"].notna() & (df["value"] > 0)]
        out = df[["code", "country", "continent", "year", "value"]].copy()
        out = out.sort_values(["code", "year"])
        save("mpi.csv", out)
    except Exception as e:
        report("mpi.csv", pd.DataFrame(), str(e))


# ── Main ──────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    # Remove stale aggregated files
    for stale in ("bubble.csv", "trends.csv", "child_labor_trends.csv",
                  "child_marriage_trends.csv", "out_of_school.csv",
                  "missing_data_modal.csv", "missing_data_countries.csv",
                  "migration_continent.csv", "migration_country.csv"):
        p = os.path.join(OUT, stale)
        if os.path.exists(p):
            os.remove(p)
            print(f"  [DEL] {stale}")

    print("\nVeni Vidi Viz -- Preprocessing\n" + "=" * 55)
    make_life_expectancy()
    make_edu_spending()
    make_edu_completion()
    make_literacy()
    make_income()
    make_population()
    make_child_labor()
    make_child_marriage()
    make_out_of_school()
    make_poverty()
    make_gini()
    make_gpi_secondary()

    make_child_mortality()
    make_maternal_mortality()
    make_remittances()
    make_migration()
    make_mpi()
    harmonize_all_processed_country_names()
    make_missing_data_registry()
    print("\nDone -> src/datasets/processed/")
    for f in sorted(os.listdir(OUT)):
        size = os.path.getsize(os.path.join(OUT, f))
        print(f"  {f:<50} {size // 1024:>4} KB")
